# -*- coding: utf-8 -*-
"""
@Time ： 2020/4/6 12:50
@Auth ： Suk
@File ： picOutput.py
@IDE  ： PyCharm
@Motto： Knowing your ignorance is the best way to succeed.
@Desc:  通过腾讯文档发起的图片收集表导出图片
"""
import os
import openpyxl
import requests

root = "dist/"
wb = openpyxl.load_workbook('测试.xlsx')
ws = wb.get_sheet_by_name('Sheet2')
stuNum = 1  # 填写人数
pic1Colum = 4 # 图片1的超链接所在列数
pic2Colum = 10 # 图片2的超链接所在列数
picTitle1 = "二次审查查询漫游地.jpg"  # 图片1文件名：  姓名-二次审查查询漫游地.jpg
picTitle2 = "二次审查苏康码-绿色.jpg"  # 图片2文件名：  姓名-二次审查苏康码-绿色.jpg

r = 2       # 数据行首行
stuList = []
picList1 = []
picList2 = []

for r in range(2, 2 + stuNum):
    c = 3
    stuList.append(ws.cell(row=r, column=c).value)
print(stuList)

for r in range(2, 2 + stuNum):
    picList1.append(dict(ws.cell(row=r, column=pic1Colum).hyperlink)["display"])
print(picList1)

for r in range(2, 2 + stuNum):
    picList2.append(dict(ws.cell(row=r, column=pic2Colum).hyperlink)["display"])
print(picList2)

for i in range(len(stuList)):
    url = picList1[i]
    url1 = picList2[i]

    r = requests.get(url)
    r1 = requests.get(url1)

    if not os.path.exists(root + stuList[i] + "/"):
        os.mkdir(root + stuList[i] + "/")

    with open(root + stuList[i] + "/" + stuList[i] + "-" + picTitle1, "wb") as f:
        f.write(r.content)
        f.close()
    with open(root + stuList[i] + "/" + stuList[i] + "-" + picTitle2, "wb") as f:
        f.write(r1.content)
        f.close()
        print(stuList[i])


def judge(path1):
    try:
        if not os.path.exists(path1):
            os.mkdir(path1)
    except:
        print("ERROR!")
